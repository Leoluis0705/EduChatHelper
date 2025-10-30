import os
from pathlib import Path
from typing import List, Tuple, cast
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

ROOT = Path(__file__).resolve().parent.parent

# 获取考试名称和老师账号，设置考试特定的输入输出路径
exam_name = os.environ.get('EXAM_NAME', '').strip()
teacher_username = os.environ.get('TEACHER_USERNAME', '').strip()

if exam_name:
    # 创建安全的文件夹名称（移除非法字符）
    safe_exam_name = "".join(ch for ch in exam_name if ch not in '\\/:*?"<>|').strip()
    if not safe_exam_name:
        safe_exam_name = "未命名考试"
    
    # 如果有老师账号，添加到文件夹名称
    if teacher_username:
        safe_teacher_name = "".join(ch for ch in teacher_username if ch not in '\\/:*?"<>|').strip()
        if safe_teacher_name:
            safe_exam_name = f"{safe_exam_name}_{safe_teacher_name}"
    
    INPUT_XLSX = Path("outputs") / safe_exam_name / "output.xlsx"
    OUTPUT_XLSX = Path("outputs") / safe_exam_name / "output_processed.xlsx"
else:
    INPUT_XLSX = Path("outputs") / "output.xlsx"
    OUTPUT_XLSX = Path("outputs") / "output_processed.xlsx"

HEADER_MINE = "我的原文"
HEADER_MORE = "更多表达"

# 允许识别两种冒号：半角: 与全角：
COLONS = (":", "：")


def normalize_cell_text(val) -> str:
    if val is None:
        return ""
    text = str(val).replace("\r\n", "\n").replace("\r", "\n")
    # 去掉每行右侧空白，保留原始换行结构
    lines = [l.rstrip() for l in text.split("\n")]
    # 去除整体首尾空行
    return "\n".join(lines).strip()


def is_line_colon_start(line: str) -> bool:
    s = line.lstrip()
    return len(s) > 0 and s[0] in COLONS


def first_four_chars(line: str) -> str:
    # 依据“最开头四个字”——忽略前导空白
    s = line.lstrip()
    return s[:4]


def insert_above_first_colon_line(lines: List[str], new_line: str) -> List[str]:
    """
    将 new_line 插入到第一条“以冒号开头”的行的上方；
    若不存在冒号开头的行，则不插入（遵循指令中的“那一行的上面”单数语义）。
    """
    for idx, ln in enumerate(lines):
        if is_line_colon_start(ln):
            return lines[:idx] + [new_line] + lines[idx:]
    return lines


def prefix_before_colon(line: str, prefix: str) -> str:
    """
    在该行的第一个冒号前插入 prefix。
    例如: "： 内容" -> "更多表达： 内容"
    保留行的原有前导空白与冒号后内容。
    """
    leading_ws = ""
    i = 0
    while i < len(line) and line[i].isspace():
        leading_ws += line[i]
        i += 1
    # 从第一个非空白开始，若是冒号则插入
    if i < len(line) and line[i] in COLONS:
        return f"{leading_ws}{prefix}{line[i:]}"

    # 若不以冒号开头，原样返回（安全兜底）
    return line


def is_labeled_colon_line(line: str) -> bool:
    """
    判断该行是否已经带有“我的原文”或“更多表达”标签，且紧随其后为冒号。
    """
    s = line.lstrip()
    for h in (HEADER_MINE, HEADER_MORE):
        if s.startswith(h) and len(s) > len(h) and s[len(h)] in COLONS:
            return True
    return False


def process_row(mine_text: str, more_text: str) -> Tuple[str, str]:
    """
    按规则处理单行记录：
    1) 若“更多表达”中存在以冒号开头的行，且“我的原文”的最后一行以冒号开头，
       则剪切“我的原文”最后一行并插入到“更多表达”中第一条冒号行上方。
    2) 重新处理“更多表达”单元格：对其中每一条以冒号开头的行，
       查看其上一行的最开头四个字：
         - 若上一行是“我的原文”，则在该行冒号前插入“更多表达”
         - 若上一行是“更多表达”，则在该行冒号前插入“我的原文”
    """
    mine_lines = [ln for ln in normalize_cell_text(mine_text).split("\n") if ln != ""]
    more_lines = [ln for ln in normalize_cell_text(more_text).split("\n") if ln != ""]

    # 步骤 1：剪切并插入（同时为两行加上对应标签）
    has_colon_in_more = any(is_line_colon_start(ln) for ln in more_lines)
    if has_colon_in_more and mine_lines:
        last_line = mine_lines[-1]
        if is_line_colon_start(last_line):
            # 剪切“我的原文”最后一行
            cut_line = mine_lines.pop()
            # 找到“更多表达”中第一条以冒号开头的行索引
            first_colon_idx = None
            for i, ln in enumerate(more_lines):
                if is_line_colon_start(ln):
                    first_colon_idx = i
                    break
            if first_colon_idx is not None:
                # 为被剪切行添加“我的原文”标签（若尚未带标签）
                if not is_labeled_colon_line(cut_line):
                    cut_line = prefix_before_colon(cut_line, HEADER_MINE)
                # 为原第一条冒号行添加“更多表达”标签（若尚未带标签）
                target_ln = more_lines[first_colon_idx]
                if not is_labeled_colon_line(target_ln):
                    target_ln = prefix_before_colon(target_ln, HEADER_MORE)
                # 重新插入
                more_lines = more_lines[:first_colon_idx] + [cut_line] + [target_ln] + more_lines[first_colon_idx + 1:]
            else:
                # 理论不应到此分支（有冒号行才会进入），安全兜底：直接追加剪切行
                more_lines.append(cut_line)

    # 步骤 2：为“更多表达”中其他冒号行按上一行内容确定前缀（跳过已带标签的行）
    updated_more: List[str] = []
    for idx, ln in enumerate(more_lines):
        if is_line_colon_start(ln) and not is_labeled_colon_line(ln):
            # 读取上一行（若存在）
            if idx - 1 >= 0:
                prev = more_lines[idx - 1]
                head4 = first_four_chars(prev)
                if head4 == HEADER_MINE:
                    ln = prefix_before_colon(ln, HEADER_MORE)
                elif head4 == HEADER_MORE:
                    ln = prefix_before_colon(ln, HEADER_MINE)
                # 其他头部则不插入前缀
        updated_more.append(ln)

    return ("\n".join(mine_lines).strip(), "\n".join(updated_more).strip())


def main():
    if not INPUT_XLSX.exists():
        print(f"未找到 Excel 文件：{INPUT_XLSX}")
        return

    wb = load_workbook(INPUT_XLSX)
    # 默认首个工作表，名称在生成脚本中为“数据”
    ws = wb.active
    if ws is None:
        print("未找到有效工作表")
        return

    # 找到列索引
    header_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        header_row = next(header_iter)
    except StopIteration:
        print("工作表为空，缺少表头行")
        return

    header_to_index = {str(name): idx for idx, name in enumerate(header_row)}  # 0-based

    if HEADER_MINE not in header_to_index or HEADER_MORE not in header_to_index:
        print("未找到所需列：'我的原文' 或 '更多表达'")
        return

    mine_col_idx = header_to_index[HEADER_MINE] + 1  # openpyxl 使用 1-based 列号
    more_col_idx = header_to_index[HEADER_MORE] + 1

    # 辅助：若目标是合并单元格，则返回该合并区域左上角单元格进行写入
    def get_writable_cell(row_idx: int, col_idx: int) -> Cell:
        c = ws.cell(row=row_idx, column=col_idx)
        if isinstance(c, MergedCell):
            coord = c.coordinate
            for rng in ws.merged_cells.ranges:
                if coord in rng:
                    master = ws.cell(row=rng.min_row, column=rng.min_col)
                    # 类型断言：合并区域左上角应为可写的普通 Cell
                    return cast(Cell, master)
        return cast(Cell, c)

    # 遍历数据行（从第2行起）
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        mine_cell_read = row[mine_col_idx - 1]
        more_cell_read = row[more_col_idx - 1]

        new_mine, new_more = process_row(
            str(mine_cell_read.value) if mine_cell_read.value is not None else "",
            str(more_cell_read.value) if more_cell_read.value is not None else ""
        )

        mine_cell_write = cast(Cell, get_writable_cell(r_idx, mine_col_idx))
        more_cell_write = cast(Cell, get_writable_cell(r_idx, more_col_idx))

        # 运行时兜底保护：若异常出现合并单元格，跳过写入该行该列
        if not isinstance(mine_cell_write, MergedCell):
            mine_cell_write.value = new_mine
        if not isinstance(more_cell_write, MergedCell):
            more_cell_write.value = new_more

    # 输出到新文件，避免覆盖原始文件
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"处理完成，已保存：{OUTPUT_XLSX}")


if __name__ == "__main__":
    main()